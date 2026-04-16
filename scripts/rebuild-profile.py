"""
每日档案更新脚本
由每日23点总结cron调用
功能：读取skill-usage.json，将使用次数写入档案的技能一览sheet
用法：python scripts/rebuild-profile.py
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
from pathlib import Path

OUTPUT = r"C:\Users\Administrator\Desktop\贾维斯个人档案.xlsx"
SKILL_USAGE_FILE = Path(__file__).parent.parent / "memory" / "skill-usage.json"
NOW = datetime.now(timezone(timedelta(hours=8)))

DARK = "1a1a2e"
PURPLE = "667eea"
GRAY_BG = "f7fafc"
GRAY_TEXT = "6b7280"

def hdr(cell, text, bg=DARK):
    cell.value = text
    cell.font = Font(bold=True, color="ffffff", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def title(cell, text, bg=PURPLE):
    cell.value = text
    cell.font = Font(bold=True, color="ffffff", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def dat(cell, val, bold=False, bg=None, fg="1f2937", wrap=False, align="left", size=10):
    cell.value = val
    cell.font = Font(bold=bold, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

def status_cell(cell, raw):
    if raw == "ok":
        color, bg, text = "16691a", "dcfce7", "正常"
    elif raw == "error":
        color, bg, text = "92400e", "fef3c7", "异常待修"
    elif raw == "installed":
        color, bg, text = "16691a", "dcfce7", "已安装"
    elif raw == "used":
        color, bg, text = "16691a", "dcfce7", "已使用"
    elif raw == "wait":
        color, bg, text = "6b7280", "f3f4f6", "待激活"
    elif raw == "not_installed":
        color, bg, text = "991b1b", "fee2e2", "未安装"
    else:
        color, bg, text = "6b7280", "f3f4f6", str(raw)
    cell.value = text
    cell.font = Font(bold=True, color=color, size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def can_use(cell, raw):
    if raw == "ok":
        color, bg, text = "16691a", "dcfce7", "可以正常使用"
    elif raw == "fix":
        color, bg, text = "92400e", "fef3c7", "需要修复"
    elif raw == "wait":
        color, bg, text = "6b7280", "f3f4f6", "待激活"
    elif raw == "no":
        color, bg, text = "991b1b", "fee2e2", "不可用"
    else:
        color, bg, text = "6b7280", "f3f4f6", str(raw)
    cell.value = text
    cell.font = Font(bold=True, color=color, size=9)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def border(cell):
    thin = Side(style="thin", color="e5e7eb")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def get_skill_usage():
    """读取skill-usage.json，返回{skill: count}"""
    try:
        with open(SKILL_USAGE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        from collections import Counter
        counts = Counter(r["skill"] for r in data.get("records", []))
        return counts
    except:
        return {}

def build_profile(usage_counts):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ===== Sheet 1: 基本信息 =====
    ws1 = wb.create_sheet("基本信息")
    ws1.row_dimensions[1].height = 28
    for c, h in enumerate(["项目", "内容"], 1):
        hdr(ws1.cell(1, c), h)
    cw(ws1, 1, 18); cw(ws1, 2, 65)

    rows1 = [
        ("名字", "贾维斯"),
        ("角色", "AI私人助理"),
        ("主人", "李凌志"),
        ("时区", "Asia/Shanghai (GMT+8)"),
        ("激活日期", "2026-04-05"),
        ("当前版本", "1.0.9"),
        ("座右铭", "主动预判，持续进化，越来越懂你。"),
        ("Emoji", "🧠"),
        ("GitHub账号", "jimmqw（morai.top + novelpick.top）"),
        ("百度统计ID", "d1d9d04b764a3f8f5a92e975825446e6"),
        ("Tavily API Key", "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"),
        ("饿饭加速器端口", "7897（HTTP代理）"),
        ("沟通风格", "专业简洁、靠谱高效、贴心主动"),
        ("偏好规则", "做完必须主动总结（2026-04-08修订）；做完不沉默，立即汇报"),
        ("最后更新时间", NOW.strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (k, v) in enumerate(rows1, 2):
        ws1.row_dimensions[i].height = 22
        dat(ws1.cell(i, 1), k, bold=True, bg=GRAY_BG)
        dat(ws1.cell(i, 2), v, wrap=True)
        border(ws1.cell(i, 1)); border(ws1.cell(i, 2))

    # ===== Sheet 2: 版本历史 =====
    ws2 = wb.create_sheet("版本历史")
    ws2.row_dimensions[1].height = 28
    for c, h in enumerate(["版本", "日期", "主要变化", "来源"], 1):
        hdr(ws2.cell(1, c), h)
    cw(ws2, 1, 9); cw(ws2, 2, 13); cw(ws2, 3, 72); cw(ws2, 4, 22)

    versions = [
        ("1.0.1","2026-04-05","初始激活：名字/风格/SOUL/IDENTITY/USER建立；记忆分类V1；主人偏好记录","激活"),
        ("1.0.2","2026-04-05","Skill注册模式；提议-确认模式；Feature Flag架构；buildTool()设计","Claude Code源码第一轮"),
        ("1.0.3","2026-04-05","extractMemories服务；禁止保存列表；记忆漂移处理","Claude Code源码第三轮"),
        ("1.0.4","2026-04-05","context.ts；工具安全模式；任务状态机；快捷键系统","Claude Code源码第四轮"),
        ("1.0.5","2026-04-05","事件日志系统；GrowthBook A/B测试；插件架构；Fuse.js模糊搜索","Claude Code源码第五轮"),
        ("1.0.6","2026-04-05","Feature Flag/Forked Agent/权限系统全面总结；已应用记忆V2+禁止规则","Claude Code源码全面"),
        ("1.0.7","2026-04-08","网站67页修复+sitemap+相关文章；写作SOP+BLUF；会话恢复/MEMORY行数/记忆分类/两步保存/对抗自检/错误类","全天工作+源码第二轮"),
        ("1.0.8","2026-04-08","SessionMemory cron(每2h)；做梦cron(每3天7:00)；跨会话恢复；文件管理体系；桌面档案建立","主人要求补足三层记忆"),
        ("1.0.9","2026-04-12","技能全面更新（self-improving-agent→v3.0.13/desktop-control新装）；SEO技术修复（sitemap去重/og:image）；fateandmethod.com修复（删文章页仅留首页/gitattributes截断问题）；三个网站完整URL档案写入桌面Excel；档案每日cron整合","全天工作"),
    ]
    for i, row in enumerate(versions, 2):
        ws2.row_dimensions[i].height = 26
        for c, v in enumerate(row, 1):
            dat(ws2.cell(i, c), v, wrap=True)
            border(ws2.cell(i, c))

    # ===== Sheet 3: 技能一览(55) =====
    ws3 = wb.create_sheet("技能一览(55)")
    ws3.row_dimensions[1].height = 30
    for c, h in enumerate(["编号", "技能名称", "技能作用/用途", "安装状态", "使用状态", "可否正常使用", "使用次数", "备注"], 1):
        title(ws3.cell(1, c), h)
    cw(ws3, 1, 8); cw(ws3, 2, 24); cw(ws3, 3, 44); cw(ws3, 4, 14); cw(ws3, 5, 14); cw(ws3, 6, 16); cw(ws3, 7, 10); cw(ws3, 8, 30)

    SKILLS = [
        ("001","clawhub","搜索/安装/更新/发布技能（clawhub.com）","installed","used","ok",""),
        ("002","find-skills","搜索并安装新技能，支持clawhub和skillhub","installed","used","ok",""),
        ("003","skillhub-preference","优先使用skillhub商店，fallback到clawhub","installed","used","ok",""),
        ("004","clawhub-preference","优先使用clawhub商店","installed","used","ok",""),
        ("005","github","gh CLI操作GitHub：issues/PRs/CI/代码审查/git push","installed","used","ok",""),
        ("006","multi-search-engine","17个搜索引擎集成（8个CN+9个全球）","installed","used","ok",""),
        ("007","weather","获取天气预报，无需API key","installed","unused","ok",""),
        ("008","clawvard_exam","参加Clawvard入学考试（16题AI评分）","installed","used","ok","得分87.5/A级"),
        ("009","feishu-doc","飞书文档读写操作","installed","error","fix","cron投递到飞书失败"),
        ("010","feishu-drive","飞书云盘文件管理","installed","error","fix","同上"),
        ("011","feishu-perm","飞书文档和文件权限管理","installed","error","fix","同上"),
        ("012","feishu-wiki","飞书知识库导航","installed","error","fix","同上"),
        ("013","memory-hygiene","清理优化LanceDB向量记忆，防止内存膨胀","installed","unused","ok",""),
        ("014","memory-tiering","自动化多层级记忆管理（HOT/WARM/COLD）","installed","unused","ok",""),
        ("015","openviking-memory","OpenViking长期记忆插件，自动记忆持久化","installed","unused","wait","未安装插件本体"),
        ("016","self-improving-agent","捕获教训/错误/纠正，持续改进","installed","unused","ok",""),
        ("017","proactive-agent-lite","主动型代理，从被动执行转为主动伙伴","installed","unused","ok",""),
        ("018","careful","破坏性命令警告和安全检查","installed","unused","ok",""),
        ("019","review","PR预审，检查范围漂移，强化需求理解","installed","unused","ok",""),
        ("020","office-hours","YC风格产品需求分析，强制深入理解再动手","installed","unused","ok",""),
        ("021","humanizer","去除AI写作痕迹（膨胀符号/promo语言等）","installed","unused","ok",""),
        ("022","skill-vetter","技能安全审核，安装前检查恶意模式","installed","unused","ok",""),
        ("023","website-seo","网站SEO完整系统（页面优化/schema/技术清单）","installed","unused","ok",""),
        ("024","seo-optimization","SEO优化可落地指南与SOP","installed","unused","ok",""),
        ("025","tavily-search","AI优化搜索（Tavily API）","installed","unused","ok",""),
        ("026","openclaw-tavily-search","OpenClaw Tavily搜索集成","installed","unused","ok",""),
        ("027","agent-browser","Rust headless浏览器自动化CLI","installed","used","ok",""),
        ("028","context-compressor","会话上下文压缩工具","installed","unused","ok",""),
        ("029","copywriter","UX文案、营销内容、产品文案撰写","installed","unused","ok",""),
        ("030","ai-seo-writer","SEO优化博客文章写作","installed","unused","ok",""),
        ("031","elite-frontend-design","前端UI界面设计（HTML/CSS/JS）","installed","unused","ok",""),
        ("032","gh-issues","抓取GitHub issues，spawn子agent实现修复并开PR","installed","unused","ok",""),
        ("033","node-connect","诊断OpenClaw节点连接和配对失败","installed","unused","ok",""),
        ("034","healthcheck","主机安全加固和风险配置检查","installed","unused","ok",""),
        ("035","skill-creator","创建/编辑/审计AgentSkills技能","installed","unused","ok",""),
        ("036","gog","（用途未知，需读取SKILL.md确认）","not_installed","unused","no",""),
        ("037","image-gen","图片生成技能","installed","unused","ok",""),
        # 课程附赠技能
        ("038","192-battery-customer-research","电池行业客户信息收集与整理（锂电池/铅酸电池/储能电池）","installed","unused","ok",""),
        ("039","193-bid-document-creator","标书文档创作与解析（PDF/图片/表格）","installed","unused","ok",""),
        ("040","194-capcut-video-editor","视频剪辑与营销视频制作（CapCut/剪映）","installed","unused","ok",""),
        ("041","196-comfyui-painter","本地ComfyUI画图+CivitAI模型搜索/下载","installed","unused","ok",""),
        ("042","199-douyin-hot-products","抖音热卖商品调研与营销规划","installed","unused","ok",""),
        ("043","200-douyin-publisher","自动发布视频到抖音","installed","unused","ok",""),
        ("044","201-env-equip-customer-research","环保设备行业客户调研与分析","installed","unused","ok",""),
        ("045","205-movie-commentary","影视解说文案创作与视频剪辑","installed","unused","ok",""),
        ("046","206-multi-search-engine","多搜索引擎集成（8CN+9全球）","installed","used","ok","同序号006"),
        ("047","209-douyin-video-parse","抖音URL解析+视频直链获取+下载","installed","unused","ok",""),
        ("048","211-shortdrama-script","短剧剧本创作（霸总/逆袭/复仇等）","installed","unused","ok",""),
        ("049","213-tech-news-daily","每日科技市场动态/AI技术新闻","installed","unused","ok",""),
        ("050","215-agent-complete","补全OpenClaw Agent完整配置","installed","unused","ok",""),
        ("051","219-intl-news","国际新闻抓取与整理（多源）","installed","unused","ok",""),
        ("052","216","（课程附赠，待读取SKILL.md确认）","installed","unused","wait",""),
        ("053","217","（课程附赠，待读取SKILL.md确认）","installed","unused","wait",""),
        ("054","218","（课程附赠，待读取SKILL.md确认）","installed","unused","wait",""),
    ]

    for i, row in enumerate(SKILLS, 2):
        ws3.row_dimensions[i].height = 22
        ws3.cell(i, 1).value = row[0]
        ws3.cell(i, 1).font = Font(size=9, color=GRAY_TEXT)
        ws3.cell(i, 1).alignment = Alignment(horizontal="center", vertical="center")
        dat(ws3.cell(i, 2), row[1], bold=True)
        dat(ws3.cell(i, 3), row[2], wrap=True, fg=GRAY_TEXT, size=9)
        status_cell(ws3.cell(i, 4), row[3])
        status_cell(ws3.cell(i, 5), row[4])
        can_use(ws3.cell(i, 6), row[5])
        # 使用次数
        count = usage_counts.get(row[1], 0)
        count_color = "16691a" if count > 0 else GRAY_TEXT
        ws3.cell(i, 7).value = count if count > 0 else "-"
        ws3.cell(i, 7).font = Font(bold=(count > 0), color=count_color, size=10)
        ws3.cell(i, 7).alignment = Alignment(horizontal="center", vertical="center")
        dat(ws3.cell(i, 8), row[6], fg=GRAY_TEXT, size=9, wrap=True)
        for c in range(1, 9):
            border(ws3.cell(i, c))

    # ===== Sheet 4: 能力体系 =====
    ws4 = wb.create_sheet("能力体系")
    ws4.row_dimensions[1].height = 30
    for c, h in enumerate(["类别", "能力", "状态", "可否正常使用", "备注"], 1):
        title(ws4.cell(1, c), h)
    cw(ws4, 1, 18); cw(ws4, 2, 30); cw(ws4, 3, 16); cw(ws4, 4, 16); cw(ws4, 5, 32)

    CAPS = [
        ("记忆-L1自动提取","SessionMemory cron每2小时","ok","ok",""),
        ("记忆-L2长期索引","MEMORY.md长期记忆","ok","ok",""),
        ("记忆-L3定期归档","做梦机制cron每3天7:00","ok","ok",""),
        ("记忆-跨会话恢复","session-template+启动检查","ok","ok",""),
        ("记忆-文件管理","frontmatter+查重+双上限","ok","ok",""),
        ("记忆-技能追踪","skill-usage.json+月度分析","ok","ok","2026-04-09新增"),
        ("网站-搭建","静态GitHub Pages","ok","ok",""),
        ("网站-SEO","sitemap/og/canonical/Schema","ok","ok",""),
        ("网站-每日文章","每站每天1篇20:00","ok","ok",""),
        ("网站-GSC","Google Search Console","wait","wait","需主人做DNS验证"),
        ("网站-移动端","Lighthouse实测","error","fix",""),
        ("写作-SERP分析","写前竞品分析","ok","ok",""),
        ("写作-BLUF框架","先给结论再展开","ok","ok",""),
        ("写作-文章SOP","完整发布流程","ok","ok",""),
        ("定时-网站巡检","每天09:00","error","fix","Feishu投递问题"),
        ("定时-技能市场","每天10:00","ok","ok",""),
        ("定时-旧文章检查","每3天10:00","ok","ok",""),
        ("定时-文章发布","每天20:00","ok","ok",""),
        ("定时-每日总结","每天23:00（含档案更新）","ok","ok",""),
        ("定时-SessionMemory","每2小时","ok","ok",""),
        ("定时-做梦机制","每3天07:00","ok","ok",""),
        ("定时-每周周报","每周一09:00","ok","ok",""),
        ("定时-每月复盘","每月1号09:00","ok","ok",""),
        ("优化-错误类体系","AbortError/ShellError/ConfigError","ok","ok",""),
        ("优化-对抗自检","adversarial-check清单","ok","ok",""),
        ("优化-诊断日志","duration埋点格式","ok","ok",""),
        ("优化-状态管理","30行Store模板","ok","ok",""),
    ]

    for i, row in enumerate(CAPS, 2):
        ws4.row_dimensions[i].height = 22
        dat(ws4.cell(i, 1), row[0], bold=True)
        dat(ws4.cell(i, 2), row[1])
        status_cell(ws4.cell(i, 3), row[2])
        can_use(ws4.cell(i, 4), row[3])
        dat(ws4.cell(i, 5), row[4] or "", fg=GRAY_TEXT, size=9)
        for c in range(1, 6):
            border(ws4.cell(i, c))

    # ===== Sheet 5: 缺点与待优化 =====
    ws5 = wb.create_sheet("缺点与待优化")
    ws5.row_dimensions[1].height = 30
    for c, h in enumerate(["优先级", "缺点", "解决方案", "状态"], 1):
        title(ws5.cell(1, c), h)
    cw(ws5, 1, 10); cw(ws5, 2, 45); cw(ws5, 3, 40); cw(ws5, 4, 14)

    ISSUES = [
        ("高","网站质量巡检cron失败（Feishu投递问题）","查isolated session announce配置","fix"),
        ("高","Google Search Console未注册","等主人做DNS验证","wait"),
        ("高","移动端Lighthouse未实测","排查Lighthouse问题","fix"),
        ("中","技能注册表未建立","下次学习时顺手做","todo"),
        ("中","knowledge/*.md未加frontmatter","更新文件时顺手做","todo"),
        ("低","gog技能用途未知","读取SKILL.md确认","wait"),
        ("低","数字技能216-218用途未知","读取SKILL.md确认","wait"),
        ("低","没有MCP服务器集成","未理解协议，低优先级","no"),
    ]
    for i, row in enumerate(ISSUES, 2):
        ws5.row_dimensions[i].height = 24
        dat(ws5.cell(i, 1), row[0], bold=True, align="center")
        dat(ws5.cell(i, 2), row[1], wrap=True)
        dat(ws5.cell(i, 3), row[2], wrap=True, fg=GRAY_TEXT, size=9)
        status_cell(ws5.cell(i, 4), "ok" if row[3]=="ok" else "error" if row[3]=="fix" else "wait" if row[3]=="wait" else "no" if row[3]=="no" else "unused")
        for c in range(1, 5):
            border(ws5.cell(i, c))

    # ===== Sheet 6: 定时任务一览 =====
    ws6 = wb.create_sheet("定时任务一览")
    ws6.row_dimensions[1].height = 30
    for c, h in enumerate(["任务名称", "cron表达式", "下次执行", "运行状态", "可否正常使用"], 1):
        title(ws6.cell(1, c), h)
    cw(ws6, 1, 32); cw(ws6, 2, 18); cw(ws6, 3, 20); cw(ws6, 4, 16); cw(ws6, 5, 16)

    CRONS = [
        ("SessionMemory-会话精华自动提取","0 */2 * * *","每2小时","ok","ok"),
        ("每4小时检查session上下文","every 4h","后台运行","ok","ok"),
        ("网站质量巡检（morai-novelpick-daily-audit）","0 9 * * *","明天09:00","error","fix"),
        ("每日技能市场巡检","0 10 * * *","明天10:00","ok","ok"),
        ("旧文章检查优化","0 10 */3 * *","4月11日10:00","ok","ok"),
        ("每日文章发布（每站1篇）","0 20 * * *","明天20:00","ok","ok"),
        ("每日23点总结（含档案更新）","0 23 * * *","今晚23:00","ok","ok"),
        ("做梦机制-记忆定期整理","0 7 */3 * *","4月11日07:00","ok","ok"),
        ("每周一生成周报","0 9 * * 1","下周一09:00","ok","ok"),
        ("每月1号月度复盘","0 9 1 * *","5月1日09:00","ok","ok"),
        ("Skill-Usage-Monthly-Analysis","0 9 1 * *","5月1日09:00","ok","ok"),
    ]
    for i, row in enumerate(CRONS, 2):
        ws6.row_dimensions[i].height = 22
        dat(ws6.cell(i, 1), row[0], bold=True)
        dat(ws6.cell(i, 2), row[1], align="center")
        dat(ws6.cell(i, 3), row[2])
        status_cell(ws6.cell(i, 4), row[3])
        can_use(ws6.cell(i, 5), row[4])
        for c in range(1, 6):
            border(ws6.cell(i, c))

    # ===== Sheet 7: 知识库索引 =====
    ws7 = wb.create_sheet("知识库索引")
    ws7.row_dimensions[1].height = 30
    for c, h in enumerate(["文件路径", "用途说明"], 1):
        title(ws7.cell(1, c), h)
    cw(ws7, 1, 48); cw(ws7, 2, 60)

    KB = [
        ("knowledge/website-ops.md","网站运营SOP（文章写作/SEO检查清单/定时任务状态）"),
        ("knowledge/self-optimization.md","技术优化模板（错误类/Store/诊断日志/记忆规范）"),
        ("knowledge/adversarial-check.md","对抗性自检清单（VerificationAgent思维/边界值/并发检查）"),
        ("memory/MEMORY-SYSTEM-V2.md","记忆系统规范（含memdir体系/frontmatter/查重规则/双上限）"),
        ("memory/skill-usage.json","技能使用追踪日志（每次使用技能后追加）"),
        ("memory/session-template.md","会话恢复模板（复杂会话结束填写/新会话启动读取）"),
        ("memory/2026-04-08.md","今日日志（网站修复/文章发布/学习/优化）"),
        ("memory/2026-04-07.md","昨日日志（Clawvard考试/技能安装/网站接管准备）"),
        ("scripts/skill-usage-analyze.py","技能使用月度分析脚本（每月1号自动运行）"),
        ("scripts/skill-freq-report.py","技能使用频率排行报告（任意时刻可跑）"),
        ("scripts/skill-archive.py","技能归档工具（90天未用技能归档）"),
        ("scripts/rebuild-profile.py","档案重建脚本（每日23点cron调用）"),
        ("CLAUDE.md","自我引导手册（行为准则/自检/记忆规范/知识库索引）"),
        ("evolutions/INDEX.md","版本历史入口（所有版本索引）"),
        ("evolutions/1.0.8/VERSION.md","最新版本详情（1.0.8版本变化）"),
    ]
    for i, row in enumerate(KB, 2):
        ws7.row_dimensions[i].height = 22
        dat(ws7.cell(i, 1), row[0], fg=PURPLE)
        dat(ws7.cell(i, 2), row[1], wrap=True, fg=GRAY_TEXT, size=9)
        for c in range(1, 3):
            border(ws7.cell(i, c))

    # ===== Sheet 8: 项目进度 =====
    ws8 = wb.create_sheet("项目进度")
    ws8.row_dimensions[1].height = 30
    for c, h in enumerate(["项目", "内容"], 1):
        hdr(ws8.cell(1, c), h)
    cw(ws8, 1, 22); cw(ws8, 2, 80)

    project_rows = [
        ("项目名称", "AI导航站双站（morai.top + novelpick.top）"),
        ("建站时间", "2026-04-07接手，v1.0.7期间完成67页修复+sitemap"),
        ("托管平台", "GitHub Pages（jimmqw账号）"),
        ("仓库地址", "github.com/jimmqw/morai-website | github.com/jimmqw/novelpick-website"),
        ("本地路径", "C:\\Users\\Administrator\\.openclaw\\workspace\\morai-website\\"),
        ("本地路径2", "C:\\Users\\Administrator\\.openclaw\\workspace\\novelpick-website\\"),
        ("百度统计ID", "d1d9d04b764a3f8f5a92e975825446e6"),
        ("文章数量", "各36篇（35篇文章+1个index），共72篇"),
        ("内容主题", "morai=AI工具推荐 | novelpick=网文推荐"),
        ("文章作者", "jimmqw（GitHub账号）"),
        ("网站配色", "morai=深紫渐变header | novelpick=紫色系header"),
        ("",""),
        ("已完成的优化", ""),
        ("SEO基础", "XML Sitemap生成推送 | Open Graph | Twitter Card | canonical URL | Schema.org"),
        ("内容结构", "深紫渐变header+nav+面包屑+侧边栏220px+相关文章3条+footer+百度统计"),
        ("文章规范", "SERP分析+BLUF框架（先结论后展开）+800-1500字+阅读时间估算"),
        ("定时发布", "每站每天20:00自动发布1篇（cron）"),
        ("GitHub推送", "每次发布后自动git add/commit/push"),
        ("旧文章检查", "每3天10:00自动检查优化一批（cron）"),
        ("",""),
        ("待完成（高优先级）", ""),
        ("GSC注册", "Google Search Console未注册，需主人做DNS验证才能接管搜索可见性"),
        ("AdSense申请", "需要流量基础，当前自然流量几乎为0，暂不满足申请条件"),
        ("外链（我们发出的）", "morai.top 34条（AI工具官网链接）+ novelpick.top 69条（Wikipedia链接）= 合计103条"),
        ("反向链接（别人指向我们）", "0条，较难获取，需长期Quora/Reddit outreach"),
        ("E-E-A-T信任信号", "缺About页面、作者署名页、联系方式页，影响搜索排名"),
        ("移动端实测", "Lighthouse跑不起来，无法验证移动端体验分"),
        ("",""),
        ("待完成（中优先级）", ""),
        ("About页面", "作者署名+个人介绍，提升网站可信度"),
        ("联系我们页面", "Contact page，增加信任信号"),
        ("内链深化", "相关文章推荐区块持续优化，提升PV和停留时间"),
        ("社交分享", "添加Twitter/微博分享按钮，增加社交信号"),
        ("",""),
        ("发展方向（长期）", ""),
        ("流量目标", "先追求Google收录 -> 再提升自然搜索排名 -> 最后申请AdSense"),
        ("内容扩展", "从35篇扩展到100篇+文章，增加长尾关键词覆盖"),
        ("品牌建设", "morai.top = AI工具首选导航站 | novelpick.top = 网文读者社区"),
        ("变现路径", "AdSense（需要稳定流量）-> 付费专栏/电子书 -> AI工具推荐佣金"),
        ("技术升级", "考虑Next.js重构提升加载速度，考虑添加评论区（Giscus/Utterances）"),
        ("多语言", "英文版本（未来可做AI工具英文评测站，竞争更小）"),
    ]

    for i, (k, v) in enumerate(project_rows, 2):
        ws8.row_dimensions[i].height = 22 if v else 10
        cell_style8_k = dat(ws8.cell(i, 1), k, bold=bool(v and k), bg=GRAY_BG if k else None)
        cell_style8_v = dat(ws8.cell(i, 2), v, wrap=True, fg=GRAY_TEXT if not v else "1f2937")
        if k:
            border(ws8.cell(i, 1)); border(ws8.cell(i, 2))

    wb.save(OUTPUT)

if __name__ == "__main__":
    usage = get_skill_usage()
    build_profile(usage)
    print(f"Profile rebuilt. {len(usage)} skills with usage data. 8 sheets total.")
